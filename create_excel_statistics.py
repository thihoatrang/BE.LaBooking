"""
Script to create Excel file with unit test statistics
"""
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

BASE_PATH = Path(__file__).parent

def read_csv_data(csv_file):
    """Read data from CSV file"""
    data = []
    with open(csv_file, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            data.append(row)
    return data

def create_excel_statistics():
    """Create Excel file with test statistics"""
    # Read CSV data
    detailed_csv = BASE_PATH / "Unit_Test_Statistics_Detailed.csv"
    summary_csv = BASE_PATH / "Unit_Test_Statistics_Summary.csv"
    
    if not detailed_csv.exists() or not summary_csv.exists():
        print("CSV files not found. Please run generate_unit_tests.py first.")
        return
    
    detailed_data = read_csv_data(detailed_csv)
    summary_data = read_csv_data(summary_csv)
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Summary Sheet
    summary_sheet = wb.create_sheet("Summary", 0)
    summary_headers = ["Category", "Count", "Tests Created", "Coverage %"]
    
    # Add title
    summary_sheet.merge_cells('A1:D1')
    title_cell = summary_sheet['A1']
    title_cell.value = "UNIT TEST STATISTICS SUMMARY"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
    title_cell.font = Font(bold=True, color="FFFFFF", size=14)
    
    # Add headers
    summary_sheet.append([])
    summary_sheet.append(summary_headers)
    
    # Style headers
    for col in range(1, len(summary_headers) + 1):
        cell = summary_sheet.cell(3, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # Add summary data
    for row_data in summary_data:
        row = [
            row_data["Category"],
            int(row_data["Count"]),
            int(row_data["Tests Created"]),
            row_data["Coverage %"]
        ]
        summary_sheet.append(row)
    
    # Style data rows
    for row in range(4, 4 + len(summary_data)):
        for col in range(1, len(summary_headers) + 1):
            cell = summary_sheet.cell(row, col)
            cell.border = border
            if col == 1:
                cell.font = Font(bold=True)
            elif col in [2, 3]:
                cell.alignment = Alignment(horizontal="center")
    
    # Auto-adjust column widths
    for col in range(1, len(summary_headers) + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in summary_sheet[column]:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max(max_length + 2, 12), 50)
        summary_sheet.column_dimensions[column].width = adjusted_width
    
    # Detailed Sheet - All Components
    detailed_sheet = wb.create_sheet("All Components")
    detailed_headers = ["No", "Component Name", "Type", "Module", "Namespace", "Test File", "Status", "Test Count", "File Path"]
    
    # Add title
    detailed_sheet.merge_cells('A1:I1')
    title_cell = detailed_sheet['A1']
    title_cell.value = "DETAILED UNIT TEST STATISTICS"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
    title_cell.font = Font(bold=True, color="FFFFFF", size=14)
    
    # Add headers
    detailed_sheet.append([])
    detailed_sheet.append(detailed_headers)
    
    # Style headers
    for col in range(1, len(detailed_headers) + 1):
        cell = detailed_sheet.cell(3, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # Add detailed data
    for idx, row_data in enumerate(detailed_data, 1):
        row = [
            int(row_data["No"]),
            row_data["Component Name"],
            row_data["Type"],
            row_data["Module"],
            row_data["Namespace"],
            row_data["Test File"],
            row_data["Status"],
            int(row_data["Test Count"]),
            row_data["File Path"]
        ]
        detailed_sheet.append(row)
    
    # Style data rows with alternating colors
    fill_even = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for row in range(4, 4 + len(detailed_data)):
        for col in range(1, len(detailed_headers) + 1):
            cell = detailed_sheet.cell(row, col)
            cell.border = border
            if (row - 4) % 2 == 0:
                cell.fill = fill_even
            if col == 1:
                cell.alignment = Alignment(horizontal="center")
            elif col in [7, 8]:
                cell.alignment = Alignment(horizontal="center")
    
    # Auto-adjust column widths
    for col in range(1, len(detailed_headers) + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in detailed_sheet[column]:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max(max_length + 2, 12), 80)
        detailed_sheet.column_dimensions[column].width = adjusted_width
    
    # Create separate sheets by type
    types = {}
    for row_data in detailed_data:
        comp_type = row_data["Type"]
        if comp_type not in types:
            types[comp_type] = []
        types[comp_type].append(row_data)
    
    # Services Sheet
    if "Service" in types or "SagaService" in types:
        services_sheet = wb.create_sheet("Services")
        services = [r for r in detailed_data if r["Type"] in ["Service", "SagaService"]]
        
        services_sheet.append(["Services Unit Tests"])
        services_sheet.append([])
        services_sheet.append(detailed_headers)
        
        # Style headers
        for col in range(1, len(detailed_headers) + 1):
            cell = services_sheet.cell(3, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Add services data
        for idx, row_data in enumerate(services, 1):
            row = [
                idx,
                row_data["Component Name"],
                row_data["Type"],
                row_data["Module"],
                row_data["Namespace"],
                row_data["Test File"],
                row_data["Status"],
                int(row_data["Test Count"]),
                row_data["File Path"]
            ]
            services_sheet.append(row)
        
        # Style data rows
        for row in range(4, 4 + len(services)):
            for col in range(1, len(detailed_headers) + 1):
                cell = services_sheet.cell(row, col)
                cell.border = border
                if (row - 4) % 2 == 0:
                    cell.fill = fill_even
        
        # Auto-adjust column widths
        for col in range(1, len(detailed_headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in services_sheet[column]:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 12), 80)
            services_sheet.column_dimensions[column].width = adjusted_width
    
    # Controllers Sheet
    if "Controller" in types:
        controllers_sheet = wb.create_sheet("Controllers")
        controllers = [r for r in detailed_data if r["Type"] == "Controller"]
        
        controllers_sheet.append(["Controllers Unit Tests"])
        controllers_sheet.append([])
        controllers_sheet.append(detailed_headers)
        
        # Style headers
        for col in range(1, len(detailed_headers) + 1):
            cell = controllers_sheet.cell(3, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Add controllers data
        for idx, row_data in enumerate(controllers, 1):
            row = [
                idx,
                row_data["Component Name"],
                row_data["Type"],
                row_data["Module"],
                row_data["Namespace"],
                row_data["Test File"],
                row_data["Status"],
                int(row_data["Test Count"]),
                row_data["File Path"]
            ]
            controllers_sheet.append(row)
        
        # Style data rows
        for row in range(4, 4 + len(controllers)):
            for col in range(1, len(detailed_headers) + 1):
                cell = controllers_sheet.cell(row, col)
                cell.border = border
                if (row - 4) % 2 == 0:
                    cell.fill = fill_even
        
        # Auto-adjust column widths
        for col in range(1, len(detailed_headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in controllers_sheet[column]:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 12), 80)
            controllers_sheet.column_dimensions[column].width = adjusted_width
    
    # Repositories Sheet
    if "Repository" in types:
        repos_sheet = wb.create_sheet("Repositories")
        repos = [r for r in detailed_data if r["Type"] == "Repository"]
        
        repos_sheet.append(["Repositories Unit Tests"])
        repos_sheet.append([])
        repos_sheet.append(detailed_headers)
        
        # Style headers
        for col in range(1, len(detailed_headers) + 1):
            cell = repos_sheet.cell(3, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Add repos data
        for idx, row_data in enumerate(repos, 1):
            row = [
                idx,
                row_data["Component Name"],
                row_data["Type"],
                row_data["Module"],
                row_data["Namespace"],
                row_data["Test File"],
                row_data["Status"],
                int(row_data["Test Count"]),
                row_data["File Path"]
            ]
            repos_sheet.append(row)
        
        # Style data rows
        for row in range(4, 4 + len(repos)):
            for col in range(1, len(detailed_headers) + 1):
                cell = repos_sheet.cell(row, col)
                cell.border = border
                if (row - 4) % 2 == 0:
                    cell.fill = fill_even
        
        # Auto-adjust column widths
        for col in range(1, len(detailed_headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in repos_sheet[column]:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 12), 80)
            repos_sheet.column_dimensions[column].width = adjusted_width
    
    # Save Excel file
    excel_file = BASE_PATH / "Unit_Test_Statistics.xlsx"
    wb.save(excel_file)
    print(f"\nExcel file created successfully: {excel_file}")
    print(f"Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"\nSheets created:")
    print(f"  - Summary")
    print(f"  - All Components")
    for sheet_name in wb.sheetnames:
        if sheet_name not in ["Summary", "All Components"]:
            print(f"  - {sheet_name}")
    
    return excel_file

if __name__ == "__main__":
    create_excel_statistics()

